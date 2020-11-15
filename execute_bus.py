from bus.bus_timetable import split_timetable

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


def execute_3102():
    dest_file = "D:\\3102.xlsx"
    workbook = openpyxl.Workbook()
    # 평일
    from_start = "05:30/05:50/06:05/06:20/06:40/07:00/07:20/07:40/08:00/08:40/09:20/10:00/10:30/11:00/11:20/11:40/12:00/12:20/12:50/13:20/13:50/14:20/14:50/15:20/15:40/16:00/16:20/16:45/17:10/17:35/18:00/18:30/19:00/19:30/19:50/20:10/20:40/21:15/21:50/22:30"
    from_end = "06:50/07:15/07:40/08:05/08:30/08:55/09:10/09:30/09:50/10:10/10:50/11:30/12:00/12:30/12:50/13:10/13:30/13:50/14:20/14:50/15:20/15:50/16:20/16:50/17:10/17:30/17:50/18:15/18:40/19:05/19:30/20:00/20:30/21:00/21:20/21:40/22:10/22:45/23:20/23:50"
    timetable_start = split_timetable(from_start)
    timetable_end = split_timetable(from_end)
    
    max_len_start = 0
    for value in timetable_start.values():
        if len(value) > max_len_start:
            max_len_start = len(value)

    max_len_end = 0
    for value in timetable_end.values():
        if len(value) > max_len_end:
            max_len_end = len(value)

    hours = sorted(list(set(list(timetable_start.keys()) + list(timetable_end.keys()))))

    # 첫째줄
    worksheet = workbook.create_sheet("평일")
    worksheet.append(["3102"])
    worksheet.merge_cells(f"A1:{get_column_letter(max_len_start + max_len_end + 1)}1")

    # 둘째줄
    row = ["송산그린시티 출발"]
    row.extend([''] * (max_len_start - 1))
    row.append("행선")
    row.append("강남역 출발")
    row.extend([''] * (max_len_end - 1))
    worksheet.append(row)
    worksheet.merge_cells(f"A2:{get_column_letter(max_len_start)}2")
    worksheet.merge_cells(f"{get_column_letter(max_len_start + 2)}2:{get_column_letter(max_len_start + max_len_end + 1)}2")

    # 셋째줄 이상
    for hour in hours:
        if hour not in timetable_start.keys():
            timetable_start[hour] = []
        if hour not in timetable_end.keys():
            timetable_end[hour] = []
        row = timetable_start[hour]
        row.extend([''] * (max_len_start - len(row)))
        row.append(hour)
        row.extend(timetable_end[hour])
        row.extend([''] * (max_len_start + max_len_start + 1 - len(row)))
        worksheet.append(row)

    # 토요일
    from_start = "05:40/06:10/06:35/06:50/07:20/08:00/08:40/09:20/10:00/10:40/11:20/12:00/12:40/13:20/14:00/14:40/15:20/16:00/16:35/17:15/17:55/18:35/19:15/19:55/20:30/21:10/21:50/22:30"
    from_end = "06:45/07:40/08:10/08:30/09:00/09:30/10:10/10:50/11:30/12:10/12:50/13:30/14:05/14:45/15:25/16:05/16:45/17:25/18:00/18:40/19:20/20:00/20:40/21:20/21:55/22:35/23:10/23:40"
    timetable_start = split_timetable(from_start)
    timetable_end = split_timetable(from_end)
    
    max_len_start = 0
    for value in timetable_start.values():
        if len(value) > max_len_start:
            max_len_start = len(value)

    max_len_end = 0
    for value in timetable_end.values():
        if len(value) > max_len_end:
            max_len_end = len(value)

    hours = sorted(list(set(list(timetable_start.keys()) + list(timetable_end.keys()))))

    # 첫째줄
    worksheet = workbook.create_sheet("토요일")
    worksheet.append(["3102"])
    worksheet.merge_cells(f"A1:{get_column_letter(max_len_start + max_len_end + 1)}1")

    # 둘째줄
    row = ["송산그린시티 출발"]
    row.extend([''] * (max_len_start - 1))
    row.append("행선")
    row.append("강남역 출발")
    row.extend([''] * (max_len_end - 1))
    worksheet.append(row)
    worksheet.merge_cells(f"A2:{get_column_letter(max_len_start)}2")
    worksheet.merge_cells(f"{get_column_letter(max_len_start + 2)}2:{get_column_letter(max_len_start + max_len_end + 1)}2")

    # 셋째줄 이상
    for hour in hours:
        if hour not in timetable_start.keys():
            timetable_start[hour] = []
        if hour not in timetable_end.keys():
            timetable_end[hour] = []
        row = timetable_start[hour]
        row.extend([''] * (max_len_start - len(row)))
        row.append(hour)
        row.extend(timetable_end[hour])
        row.extend([''] * (max_len_start + max_len_start + 1 - len(row)))
        worksheet.append(row)

    # 일요일
    from_start = "05:40/06:15/06:50/07:25/08:05/08:45/09:30/10:15/11:00/11:45/12:40/13:25/14:10/14:55/15:40/16:30/17:15/18:00/18:45/19:30/20:10/20:55/21:45/22:30"
    from_end = "06:45/07:40/08:10/08:45/09:25/10:15/11:00/11:45/12:30/13:15/14:05/14:50/15:35/16:20/17:05/17:55/18:40/19:25/20:10/20:55/21:30/22:15/23:00/23:40"
    timetable_start = split_timetable(from_start)
    timetable_end = split_timetable(from_end)
    
    max_len_start = 0
    for value in timetable_start.values():
        if len(value) > max_len_start:
            max_len_start = len(value)

    max_len_end = 0
    for value in timetable_end.values():
        if len(value) > max_len_end:
            max_len_end = len(value)

    hours = sorted(list(set(list(timetable_start.keys()) + list(timetable_end.keys()))))

    # 첫째줄
    worksheet = workbook.create_sheet("일요일")
    worksheet.append(["3102"])
    worksheet.merge_cells(f"A1:{get_column_letter(max_len_start + max_len_end + 1)}1")

    # 둘째줄
    row = ["송산그린시티 출발"]
    row.extend([''] * (max_len_start - 1))
    row.append("행선")
    row.append("강남역 출발")
    row.extend([''] * (max_len_end - 1))
    worksheet.append(row)
    worksheet.merge_cells(f"A2:{get_column_letter(max_len_start)}2")
    worksheet.merge_cells(f"{get_column_letter(max_len_start + 2)}2:{get_column_letter(max_len_start + max_len_end + 1)}2")

    # 셋째줄 이상
    for hour in hours:
        if hour not in timetable_start.keys():
            timetable_start[hour] = []
        if hour not in timetable_end.keys():
            timetable_end[hour] = []
        row = timetable_start[hour]
        row.extend([''] * (max_len_start - len(row)))
        row.append(hour)
        row.extend(timetable_end[hour])
        row.extend([''] * (max_len_start + max_len_start + 1 - len(row)))
        worksheet.append(row)

    workbook.save(dest_file)
    workbook.close()


def execute_10():
    dest_file = "D:\\10.xlsx"
    workbook = openpyxl.Workbook()
    # 평일
    from_start = "05:40/05:55/06:10/06:20/06:30/06:40/06:50/07:00/07:10/07:20/07:30/07:40/07:50/08:00/08:10/08:20/08:30/08:40/08:50/09:00/09:10/09:30/09:50/10:10/10:30/10:50/11:10/11:30/11:45/12:00/12:10/12:20/12:30/12:40/12:50/13:00/13:10/13:20/13:30/13:40/13:50/14:00/14:15/14:30/14:45/15:00/15:15/15:30/15:45/16:00/16:15/16:30/16:45/17:00/17:15/17:30/17:40/17:50/18:00/18:10/18:20/18:30/18:40/18:50/19:00/19:10/19:20/19:30/19:40/19:50/20:00/20:10/20:20/20:30/20:40/20:50/21:00/21:20/21:35/21:50/22:10/22:30/22:50/23:15"
    from_end = "06:10/06:25/06:35/06:45/06:55/07:05/07:15/07:30/07:40/07:50/08:00/08:10/08:20/08:30/08:40/08:50/09:00/09:10/09:20/09:30/09:40/10:00/10:20/10:40/11:00/11:20/11:40/12:00/12:15/12:30/12:40/12:50/13:00/13:10/13:20/13:30/13:40/13:50/14:00/14:10/14:20/14:30/14:45/15:00/15:15/15:30/15:45/16:00/16:15/16:30/16:45/17:00/17:15/17:30/17:45/18:00/18:10/18:20/18:30/18:40/18:50/19:00/19:10/19:20/19:30/19:40/19:50/20:00/20:10/20:20/20:30/20:40/20:50/21:00/21:10/21:20/21:30/21:50/22:05/22:20/22:40/23:00/23:20/23:40"
    timetable_start = split_timetable(from_start)
    timetable_end = split_timetable(from_end)
    
    max_len_start = 0
    for value in timetable_start.values():
        if len(value) > max_len_start:
            max_len_start = len(value)

    max_len_end = 0
    for value in timetable_end.values():
        if len(value) > max_len_end:
            max_len_end = len(value)

    hours = sorted(list(set(list(timetable_start.keys()) + list(timetable_end.keys()))))

    # 첫째줄
    worksheet = workbook.create_sheet("평일")
    worksheet.append(["10"])
    worksheet.merge_cells(f"A1:{get_column_letter(max_len_start + max_len_end + 1)}1")

    # 둘째줄
    row = ["송산그린시티 출발"]
    row.extend([''] * (max_len_start - 1))
    row.append("행선")
    row.append("중앙역 출발")
    row.extend([''] * (max_len_end - 1))
    worksheet.append(row)
    worksheet.merge_cells(f"A2:{get_column_letter(max_len_start)}2")
    worksheet.merge_cells(f"{get_column_letter(max_len_start + 2)}2:{get_column_letter(max_len_start + max_len_end + 1)}2")

    # 셋째줄 이상
    for hour in hours:
        if hour not in timetable_start.keys():
            timetable_start[hour] = []
        if hour not in timetable_end.keys():
            timetable_end[hour] = []
        row = timetable_start[hour]
        row.extend([''] * (max_len_start - len(row)))
        row.append(hour)
        row.extend(timetable_end[hour])
        row.extend([''] * (max_len_start + max_len_start + 1 - len(row)))
        worksheet.append(row)

    # 토요일
    from_start = "05:50/06:05/06:20/06:35/06:50/07:05/07:20/07:35/07:50/08:05/08:20/08:40/09:00/09:20/09:40/10:00/10:25/10:50/11:15/11:40/12:05/12:30/12:55/13:20/13:45/14:00/14:15/14:30/14:45/15:00/15:20/15:40/16:00/16:20/16:40/17:00/17:20/17:40/18:00/18:20/18:35/18:50/19:05/19:20/19:35/19:50/20:05/20:20/20:35/20:50/21:05/21:20/21:35/21:50/22:05/22:20/22:35/22:50/23:05/23:15"
    from_end = "06:10/06:30/06:45/07:00/07:15/07:30/07:45/08:00/08:15/08:30/08:50/09:10/09:30/09:50/10:10/10:30/10:55/11:20/11:45/12:10/12:35/13:00/13:25/13:50/14:15/14:30/14:45/15:00/15:15/15:30/15:50/16:10/16:30/16:50/17:10/17:30/17:50/18:10/18:30/18:50/19:05/19:20/19:35/19:50/20:05/20:20/20:35/20:50/21:05/21:20/21:35/21:50/22:05/22:20/22:35/22:45/23:00/23:15/23:30/23:40"
    timetable_start = split_timetable(from_start)
    timetable_end = split_timetable(from_end)
    
    max_len_start = 0
    for value in timetable_start.values():
        if len(value) > max_len_start:
            max_len_start = len(value)

    max_len_end = 0
    for value in timetable_end.values():
        if len(value) > max_len_end:
            max_len_end = len(value)

    hours = sorted(list(set(list(timetable_start.keys()) + list(timetable_end.keys()))))

    # 첫째줄
    worksheet = workbook.create_sheet("토요일")
    worksheet.append(["10"])
    worksheet.merge_cells(f"A1:{get_column_letter(max_len_start + max_len_end + 1)}1")

    # 둘째줄
    row = ["송산그린시티 출발"]
    row.extend([''] * (max_len_start - 1))
    row.append("행선")
    row.append("중앙역 출발")
    row.extend([''] * (max_len_end - 1))
    worksheet.append(row)
    worksheet.merge_cells(f"A2:{get_column_letter(max_len_start)}2")
    worksheet.merge_cells(f"{get_column_letter(max_len_start + 2)}2:{get_column_letter(max_len_start + max_len_end + 1)}2")

    # 셋째줄 이상
    for hour in hours:
        if hour not in timetable_start.keys():
            timetable_start[hour] = []
        if hour not in timetable_end.keys():
            timetable_end[hour] = []
        row = timetable_start[hour]
        row.extend([''] * (max_len_start - len(row)))
        row.append(hour)
        row.extend(timetable_end[hour])
        row.extend([''] * (max_len_start + max_len_start + 1 - len(row)))
        worksheet.append(row)

    # 일요일
    from_start = "05:50/06:10/06:30/06:50/07:10/07:25/07:40/08:00/08:20/08:40/09:00/09:20/09:50/10:20/10:50/13:20/13:40/14:00/14:20/14:40/15:00/15:20/15:40/16:00/16:25/16:50/17:15/17:40/18:05/18:30/18:55/19:20/19:40/20:00/20:20/20:40/21:00/21:20/21:40/22:00/22:20/22:35/22:50/23:15"
    from_end = "06:10/06:35/07:00/07:20/07:40/07:55/08:10/08:30/08:50/09:10/09:30/09:50/10:20/10:50/11:20/11:50/12:20/12:50/13:20/13:50/14:10/14:30/14:50/15:10/15:30/15:50/16:10/16:30/16:55/17:20/17:45/18:10/18:35/19:00/19:25/19:50/20:10/20:30/20:50/21:10/21:30/21:50/22:10/22:30/22:50/23:10/23:25/23:40"
    timetable_start = split_timetable(from_start)
    timetable_end = split_timetable(from_end)
    
    max_len_start = 0
    for value in timetable_start.values():
        if len(value) > max_len_start:
            max_len_start = len(value)

    max_len_end = 0
    for value in timetable_end.values():
        if len(value) > max_len_end:
            max_len_end = len(value)

    hours = sorted(list(set(list(timetable_start.keys()) + list(timetable_end.keys()))))

    # 첫째줄
    worksheet = workbook.create_sheet("일요일")
    worksheet.append(["10"])
    worksheet.merge_cells(f"A1:{get_column_letter(max_len_start + max_len_end + 1)}1")

    # 둘째줄
    row = ["송산그린시티 출발"]
    row.extend([''] * (max_len_start - 1))
    row.append("행선")
    row.append("중앙역 출발")
    row.extend([''] * (max_len_end - 1))
    worksheet.append(row)
    worksheet.merge_cells(f"A2:{get_column_letter(max_len_start)}2")
    worksheet.merge_cells(f"{get_column_letter(max_len_start + 2)}2:{get_column_letter(max_len_start + max_len_end + 1)}2")

    # 셋째줄 이상
    for hour in hours:
        if hour not in timetable_start.keys():
            timetable_start[hour] = []
        if hour not in timetable_end.keys():
            timetable_end[hour] = []
        row = timetable_start[hour]
        row.extend([''] * (max_len_start - len(row)))
        row.append(hour)
        row.extend(timetable_end[hour])
        row.extend([''] * (max_len_start + max_len_start + 1 - len(row)))
        worksheet.append(row)

    workbook.save(dest_file)
    workbook.close()
execute_10()